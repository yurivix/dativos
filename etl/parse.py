"""Parse PGE-ES dativos XLSX files into normalized rows.

The original spreadsheets carry the time dimension as free text in the first
column (e.g. "16 de abril de 2024 a 15 de maio de 2024:"). The CKAN DataStore
drops this column on import, so we always re-parse from the XLSX.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import openpyxl

MONTHS_PT = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Matches "16 de abril de 2024 a 15 de maio de 2024" with tolerant whitespace,
# optional "de" before the month, and trailing punctuation.
PERIOD_RE = re.compile(
    r"(\d{1,2})\s*(?:de\s+)?([a-z]+)\s*(?:de\s+)?(\d{4})"
    r"\s*a\s*"
    r"(\d{1,2})\s*(?:de\s+)?([a-z]+)\s*(?:de\s+)?(\d{4})"
)


@dataclass(frozen=True)
class Row:
    period_start: date
    period_end: date
    mes_referencia: str  # YYYY-MM of the closing month (period_end.year, period_end.month)
    period_label: str
    n_solicitacoes: int
    n_analises: int
    valor_bruto: float


def _strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )


def parse_period(text: str) -> tuple[date, date] | None:
    """Parse a free-text period like '16 de abril de 2024 a 15 de maio de 2024'.

    Returns (start, end) or None if the text doesn't look like a period.
    """
    if not text:
        return None
    norm = _strip_accents(text).lower()
    m = PERIOD_RE.search(norm)
    if not m:
        return None
    d1, mo1, y1, d2, mo2, y2 = m.groups()
    if mo1 not in MONTHS_PT or mo2 not in MONTHS_PT:
        return None
    return (
        date(int(y1), MONTHS_PT[mo1], int(d1)),
        date(int(y2), MONTHS_PT[mo2], int(d2)),
    )


def _coerce_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(".", "").replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _coerce_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v) if isinstance(v, (int, float)) else float(
            str(v).replace(".", "").replace(",", ".")
        )
    except (TypeError, ValueError):
        return None


def parse_xlsx(path: Path) -> list[Row]:
    """Read a PGE dativos XLSX and yield normalized rows.

    The PGE spreadsheets have:
      - row 1: title (merged cells)
      - row 2: column headers
      - rows 3+: data, with column A = period text, B-D = the three metrics
    Rows where the period can't be parsed are skipped (covers blank rows and
    any totals row that might be added).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[Row] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for raw in ws.iter_rows(values_only=True):
            if not raw or len(raw) < 4:
                continue
            period_text, c2, c3, c4 = raw[0], raw[1], raw[2], raw[3]
            if not isinstance(period_text, str):
                continue
            parsed = parse_period(period_text)
            if not parsed:
                continue
            start, end = parsed
            n_sol = _coerce_int(c2)
            n_ana = _coerce_int(c3)
            valor = _coerce_float(c4)
            if n_sol is None or n_ana is None or valor is None:
                continue
            rows.append(
                Row(
                    period_start=start,
                    period_end=end,
                    mes_referencia=f"{end.year:04d}-{end.month:02d}",
                    period_label=period_text.strip().rstrip(":").strip(),
                    n_solicitacoes=n_sol,
                    n_analises=n_ana,
                    valor_bruto=valor,
                )
            )
    return rows


def parse_many(paths: Iterable[Path]) -> list[Row]:
    out: list[Row] = []
    for p in paths:
        out.extend(parse_xlsx(p))
    return out
