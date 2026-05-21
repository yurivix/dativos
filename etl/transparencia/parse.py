"""Parse the ES Transparência Advogados Dativos yearly XLSX files.

Schema observations across all 9 files (2018-2026):
  - Rows 1-6: government letterhead and free-text title.
  - Row 7 (or 8 for older files): real column header.
  - Subsequent rows: one row per payment.
  - Column count and names drift across years:
      * 2018-2024: 11-12 cols, includes 'VALOR INSS', no 'Conta Judicial'.
      * 2025+:     13 cols, adds 'Conta Judicial', drops 'VALOR INSS'.
  - 'CPF' arrives pre-masked (e.g. '***116817**').
  - 'Mês do Pagamento' is uppercase pt-BR text ('JANEIRO', 'MARÇO', ...).

The parser normalizes both layouts to a single Payment shape.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

import openpyxl

MONTHS_PT = {
    "janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4,
    "maio": 5, "junho": 6, "julho": 7, "agosto": 8,
    "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}

# Map header strings (lowercased, accent-stripped) → canonical field name.
HEADER_ALIASES = {
    "nome do beneficiario": "nome",
    "beneficiario": "nome",
    "cpf": "cpf",
    "processo judicial": "processo",
    "valor liquido": "valor_liquido",
    "valor inss": "valor_inss",
    "valor irrf": "valor_irrf",
    "retencao irrf": "valor_irrf",
    "valor bruto": "valor_bruto",
    "comarca": "comarca",
    "vara": "vara",
    "vara/nome": "vara_nome",
    "varanome": "vara_nome",
    "nome da vara": "vara_nome",
    "conta judicial": "conta_judicial",
    "mes do pagamento": "mes_pagamento",
}


@dataclass(frozen=True)
class Payment:
    nome: str
    cpf_mascarado: str | None
    processo: str
    valor_liquido: float | None
    valor_bruto: float
    valor_irrf: float | None
    valor_inss: float | None
    comarca: str | None
    vara: str | None
    vara_nome: str | None
    conta_judicial: str | None
    mes_pagamento: int  # 1..12
    ano: int
    source_download_id: int


def _norm(s: str) -> str:
    """Lowercase + strip accents + collapse whitespace + drop punctuation."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().strip()
    s = re.sub(r"[ \s]+", " ", s)
    s = re.sub(r"[^a-z0-9 /]", "", s)
    return s.strip()


def _find_header_row(ws) -> tuple[int, dict[int, str]] | None:
    """Locate the header row.

    Scans the first 20 rows. A row is the header iff at least 4 of its cells
    map to known canonical fields via HEADER_ALIASES.
    Returns (1-based row index, {column_index: canonical_field}).
    """
    for r_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=20), start=1):
        col_map: dict[int, str] = {}
        for c_idx, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            canon = HEADER_ALIASES.get(_norm(cell))
            if canon:
                col_map[c_idx] = canon
        if len(col_map) >= 4 and "nome" in col_map.values() and "processo" in col_map.values():
            return r_idx, col_map
    return None


def _parse_month(text) -> int | None:
    if not isinstance(text, str):
        return None
    norm = _norm(text)
    # Handle "MARÇO" (norm → "marco"), "DEZEMBRO/2024" etc.
    for key, num in MONTHS_PT.items():
        if norm.startswith(key):
            return num
    return None


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    # BR format: "1.234,56" → "1234.56"; tolerate "1234.56" too.
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _str_or_none(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_xlsx(path: Path, ano: int, download_id: int) -> Iterator[Payment]:
    """Stream Payment rows from one yearly Acumulado XLSX.

    `ano` is provided by the caller (the parent context knows which year this
    file represents). `download_id` is recorded as data lineage.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = _find_header_row(ws)
            if header is None:
                continue
            header_row, col_map = header
            for raw in ws.iter_rows(values_only=True, min_row=header_row + 1):
                fields: dict[str, object] = {c: None for c in HEADER_ALIASES.values()}
                for c_idx, canon in col_map.items():
                    if c_idx < len(raw):
                        fields[canon] = raw[c_idx]

                nome = _str_or_none(fields.get("nome"))
                processo = _str_or_none(fields.get("processo"))
                if not nome or not processo:
                    continue
                valor_bruto = _to_float(fields.get("valor_bruto"))
                if valor_bruto is None or valor_bruto == 0:
                    continue
                mes = _parse_month(fields.get("mes_pagamento"))
                if mes is None:
                    continue

                yield Payment(
                    nome=nome.strip(),
                    cpf_mascarado=_str_or_none(fields.get("cpf")),
                    processo=processo.strip(),
                    valor_liquido=_to_float(fields.get("valor_liquido")),
                    valor_bruto=valor_bruto,
                    valor_irrf=_to_float(fields.get("valor_irrf")),
                    valor_inss=_to_float(fields.get("valor_inss")),
                    comarca=_str_or_none(fields.get("comarca")),
                    vara=_str_or_none(fields.get("vara")),
                    vara_nome=_str_or_none(fields.get("vara_nome")),
                    conta_judicial=_str_or_none(fields.get("conta_judicial")),
                    mes_pagamento=mes,
                    ano=ano,
                    source_download_id=download_id,
                )
            break  # first sheet matched; ignore others (e.g. blank 'Planilha1')
    finally:
        wb.close()
